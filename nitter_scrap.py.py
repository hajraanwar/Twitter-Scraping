from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.styles import Font
from selenium.common.exceptions import TimeoutException

chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("detach", True)

driver = webdriver.Chrome(options=chrome_options)
driver.get("https://nitter.net/search?f=tweets")

# Import the target words list from list.py
from list import target_words

# Create an Excel workbook and add a sheet
wb = openpyxl.Workbook()
sheet = wb.active

# Add title cells
sheet.cell(row=1, column=1, value="Keywords")
sheet.cell(row=1, column=2, value="Tweet Body")
sheet.cell(row=1, column=3, value="Comments")
sheet.cell(row=1, column=4, value="Retweets")
sheet.cell(row=1, column=5, value="Quotes")
sheet.cell(row=1, column=6, value="Likes")
sheet.cell(row=1, column=1).font = Font(bold=True)
sheet.cell(row=1, column=2).font = Font(bold=True)
sheet.cell(row=1, column=3).font = Font(bold=True)
sheet.cell(row=1, column=4).font = Font(bold=True)
sheet.cell(row=1, column=5).font = Font(bold=True)
sheet.cell(row=1, column=6).font = Font(bold=True)

# Iterate over each target word
for word in target_words:
    sleep(3)
    try:
        # Clear the search box before entering a new word
        search_box = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "q"))
        )
        search_box.clear()
        search_box.send_keys(word)
        search_box.send_keys(Keys.ENTER)

        # Wait for timeline items to be present
        timeline_items = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, 'timeline-item'))
        )

        for timeline_item in timeline_items:
            sleep(3)
            try:
                # Extract information from each timeline item
                tweet_body = timeline_item.find_element(By.CLASS_NAME, 'tweet-content.media-body').text

                # Get values for 'icon-comment', 'icon-retweet', 'icon-quote', 'icon-heart'
                stats_elements = timeline_item.find_elements(By.CSS_SELECTOR, '.tweet-stats .icon-container')

                # Check if there are at least 4 elements in stats_elements before accessing them
                if len(stats_elements) >= 4:
                    comments_value = stats_elements[0].text.strip().split()[-1]
                    retweets_value = stats_elements[1].text.strip().split()[-1]
                    quotes_value = stats_elements[2].text.strip().split()[-1]
                    likes_value = stats_elements[3].text.strip().split()[-1]
                else:
                    # If there are not enough elements, assign default values or handle it accordingly
                    comments_value = "N/A"
                    retweets_value = "N/A"
                    quotes_value = "N/A"
                    likes_value = "N/A"

                # Write to Excel sheet
                sheet.append([word, tweet_body, comments_value, retweets_value, quotes_value, likes_value])

            except Exception as e:
                print(f"Error processing timeline item: {e}")

    except TimeoutException:
        print(f"TimeoutException: Timed out waiting for elements for the word '{word}'")

# Save the Excel workbook
wb.save("tweets.xlsx")

# Close the browser window
driver.quit()
